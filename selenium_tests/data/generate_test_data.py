import os
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generate_test_cases():
    data_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(data_dir, "test_data.xlsx")

    # Create a new workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Set up Headers
    headers = [
        "Test Case ID", "Module", "Test Case Name", "Description", 
        "Steps", "Expected Result", "Input Data"
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.row_dimensions[1].height = 28

    cases = []

    # 1. Login/Auth: 60 cases (TC001 to TC060)
    for i in range(1, 61):
        case_id = f"TC{i:03d}"
        module = "Login/Auth"
        
        # We vary the scenarios
        if i == 1:
            name = "Successful Login"
            desc = "Verify login works with a valid registered email and password."
            steps = "1. Navigate to /login\n2. Enter valid registered email\n3. Enter valid password\n4. Click Sign In"
            expected = "Successfully logged in and redirected to dashboard page"
            input_data = {"email": "valid_user@example.com", "password": "password123", "action": "login_success"}
        elif i <= 15:
            name = f"Login with Invalid Email Format - Variation {i-1}"
            desc = "Verify login shows error when email format is invalid."
            steps = f"1. Navigate to /login\n2. Enter invalid email variation {i-1}\n3. Enter valid password\n4. Click Sign In"
            expected = "Validation message displayed: Email format is invalid or HTML5 validation"
            emails = ["no_at_sign.com", "double@@at.com", "space in@email.com", "dots..at@email.com", "@missing_local.com"]
            email = emails[(i - 2) % len(emails)]
            input_data = {"email": f"test_{i}_{email}", "password": "password123", "action": "email_format_error"}
        elif i <= 30:
            name = f"Login with Short Password - Variation {i-15}"
            desc = "Verify login shows error when password is too short."
            steps = "1. Navigate to /login\n2. Enter valid email\n3. Enter short password\n4. Click Sign In"
            expected = "Validation message displayed: Password must be at least 6 characters or Invalid credentials"
            pwd_len = (i - 16) % 5 + 1
            input_data = {"email": f"test_short_{i}@example.com", "password": "a" * pwd_len, "action": "short_password_error"}
        elif i <= 45:
            name = f"Login with Empty Fields - Variation {i-30}"
            desc = "Verify validation when required fields are empty."
            steps = "1. Navigate to /login\n2. Enter field values (one or both empty)\n3. Click Sign In"
            expected = "HTML5 form validation error shown for required fields"
            email_val = "" if i % 3 == 0 else "some@email.com"
            pass_val = "" if i % 3 == 1 else "password123"
            input_data = {"email": email_val, "password": pass_val, "action": "empty_fields_error"}
        else:
            name = f"Login SQL Injection Safety - Variation {i-45}"
            desc = "Verify application safely rejects malicious SQL injection payloads in login fields."
            steps = "1. Navigate to /login\n2. Enter SQL payload in email/password field\n3. Click Sign In"
            expected = "Invalid email or password error displayed, application remains secure"
            payloads = ["' OR 1=1 --", "' UNION SELECT NULL --", "admin'--", "' OR '1'='1"]
            payload = payloads[(i - 46) % len(payloads)]
            input_data = {"email": f"{payload}@example.com", "password": "password123", "action": "sql_injection_rejection"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # 2. Signup/Registration: 80 cases (TC061 to TC140)
    for i in range(61, 141):
        case_id = f"TC{i:03d}"
        module = "Signup/Registration"
        
        idx = i - 60
        if idx == 1:
            name = "Successful Account Registration"
            desc = "Verify new user registration works with unique email and passwords matching."
            steps = "1. Navigate to /register\n2. Enter full name\n3. Enter unique unused email\n4. Enter password\n5. Confirm password\n6. Click Create Account"
            expected = "Account successfully registered and redirected to dashboard page"
            input_data = {"name": "New User", "email": "unique_register_test@example.com", "password": "password123", "confirmPassword": "password123", "action": "register_success"}
        elif idx <= 15:
            name = f"Registration Password Mismatch - Variation {idx-1}"
            desc = "Verify validation error when password and confirm password do not match."
            steps = "1. Navigate to /register\n2. Enter name and email\n3. Enter password\n4. Enter mismatched confirm password\n5. Click Create Account"
            expected = "Error toast displayed: Passwords do not match"
            input_data = {"name": "Test User", "email": f"mismatch_{i}@example.com", "password": "password123", "confirmPassword": "different_pwd123", "action": "password_mismatch"}
        elif idx <= 35:
            name = f"Registration Short Password - Variation {idx-15}"
            desc = "Verify registration fails and shows error for password < 6 characters."
            steps = "1. Navigate to /register\n2. Enter name, email\n3. Enter short password\n4. Click Create Account"
            expected = "Error toast displayed: Password must be at least 6 characters"
            pwd_len = (idx - 16) % 5 + 1
            input_data = {"name": "Test User", "email": f"short_{i}@example.com", "password": "a" * pwd_len, "confirmPassword": "a" * pwd_len, "action": "short_password_register"}
        elif idx <= 55:
            name = f"Registration Missing Fields - Variation {idx-35}"
            desc = "Verify form validations for empty inputs on registration fields."
            steps = "1. Navigate to /register\n2. Clear fields selectively\n3. Click Create Account"
            expected = "HTML5 form validation error shown for missing required fields"
            name_val = "" if idx % 4 == 0 else "John Doe"
            email_val = "" if idx % 4 == 1 else "john@example.com"
            pass_val = "" if idx % 4 == 2 else "password123"
            conf_val = "" if idx % 4 == 3 else "password123"
            input_data = {"name": name_val, "email": email_val, "password": pass_val, "confirmPassword": conf_val, "action": "missing_fields_register"}
        elif idx <= 70:
            name = f"Registration Already Existing Email - Variation {idx-55}"
            desc = "Verify registration fails when email is already registered."
            steps = "1. Navigate to /register\n2. Enter name, existing email, and passwords\n3. Click Create Account"
            expected = "Error toast displayed: Email already registered"
            input_data = {"name": "Duplicate User", "email": "valid_user@example.com", "password": "password123", "confirmPassword": "password123", "action": "duplicate_email_register"}
        else:
            name = f"Registration Name Bounds check - Variation {idx-70}"
            desc = "Verify registration rejects invalid names or allows edge case names."
            steps = "1. Navigate to /register\n2. Enter special/empty name\n3. Click Create Account"
            expected = "Errors handled or name saved correctly, system stays stable"
            names = ["A", "Very " * 10 + "Long Name", "Name123", "Name!@#", "   Spaces   "]
            name_val = names[(idx - 71) % len(names)]
            input_data = {"name": name_val, "email": f"name_bound_{i}@example.com", "password": "password123", "confirmPassword": "password123", "action": "name_bounds_check"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # 3. Navigation/Menu: 60 cases (TC141 to TC200)
    for i in range(141, 201):
        case_id = f"TC{i:03d}"
        module = "Navigation/Menu"
        
        idx = i - 140
        if idx == 1:
            name = "Verify Modeler Link"
            desc = "Verify clicking Modeler link in header routes to /modeler page."
            steps = "1. Click Modeler link\n2. Verify URL path matches /modeler or displays Scheduled Intentions"
            expected = "Page URL contains /modeler and heading shows Scheduled Intentions"
            input_data = {"target_page": "/modeler", "expected_text": "Scheduled Intentions", "action": "nav_modeler"}
        elif idx == 2:
            name = "Verify Insights Link"
            desc = "Verify clicking Insights link in header routes to /insights page."
            steps = "1. Click Insights link\n2. Verify URL path matches /insights"
            expected = "Page URL contains /insights and charts render"
            input_data = {"target_page": "/insights", "expected_text": "Insights", "action": "nav_insights"}
        elif idx == 3:
            name = "Verify Pivot Link"
            desc = "Verify clicking Pivot link in header routes to /pivot page."
            steps = "1. Click Pivot link\n2. Verify URL path matches /pivot"
            expected = "Page URL contains /pivot and components render"
            input_data = {"target_page": "/pivot", "expected_text": "Pivot", "action": "nav_pivot"}
        elif idx == 4:
            name = "Verify Sync Link"
            desc = "Verify clicking Sync link in header routes to /sync page."
            steps = "1. Click Sync link\n2. Verify URL path matches /sync"
            expected = "Page URL contains /sync"
            input_data = {"target_page": "/sync", "expected_text": "Sync", "action": "nav_sync"}
        elif idx <= 20:
            name = f"Unauthorized Navigation Redirect - Variation {idx-4}"
            desc = "Verify unauthorized user attempting to access private pages is redirected to /login."
            steps = f"1. Clear cookies/local storage\n2. Direct navigate to private page {idx-4}\n3. Verify redirect to /login"
            expected = "Redirected to /login page with auth screen"
            private_pages = ["/modeler", "/insights", "/pivot", "/sync", "/"]
            target = private_pages[(idx - 5) % len(private_pages)]
            input_data = {"target_page": target, "action": "unauth_redirect"}
        elif idx <= 40:
            name = f"Tab Navigation Active Classes - Variation {idx-20}"
            desc = "Verify clicking menu links highlights active tab in navigation header."
            steps = "1. Navigate to target page\n2. Inspect navigation DOM\n3. Verify active class is applied to tab"
            expected = "Active style classes (e.g. text-primary) applied to current page's link"
            pages = ["/modeler", "/insights", "/pivot", "/sync"]
            target = pages[(idx - 21) % len(pages)]
            input_data = {"target_page": target, "action": "active_tab_check"}
        else:
            name = f"Footer Link Validation - Variation {idx-40}"
            desc = "Verify footer layout contains valid references or resources."
            steps = "1. Scroll to footer\n2. Check links structure or text elements"
            expected = "Footer items present and stable layout"
            input_data = {"action": "footer_validation"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # 4. Scheduled Intentions Form: 100 cases (TC201 to TC300)
    for i in range(201, 301):
        case_id = f"TC{i:03d}"
        module = "Scheduled Intentions Form"
        
        idx = i - 200
        if idx == 1:
            name = "Create Intention Successfully"
            desc = "Verify adding a valid intention creates it and lists it on page."
            steps = "1. Navigate to /modeler\n2. Enter Title, select Category, select Effort and Duration\n3. Click Create Intention button"
            expected = "Intention added successfully and appears in the Scheduled list"
            input_data = {"title": "Daily Cardio Routine", "category": "health", "effort": "3", "duration": "45", "action": "create_success"}
        elif idx <= 25:
            name = f"Create Intention Empty Title - Variation {idx-1}"
            desc = "Verify validation error when title is empty."
            steps = "1. Navigate to /modeler\n2. Leave Title empty\n3. Click Create Intention button"
            expected = "HTML5 form validation error or toast validation message for empty title"
            input_data = {"title": "", "category": "health", "effort": "2", "duration": "30", "action": "empty_title_error"}
        elif idx <= 50:
            name = f"Create Intention Duration Boundaries - Variation {idx-25}"
            desc = "Verify duration limits (zero, negative, high numbers)."
            steps = "1. Navigate to /modeler\n2. Fill Title, select category\n3. Input out-of-bound duration\n4. Click Create"
            expected = "Validation message displayed or input corrected to boundary limits"
            durations = [-5, 0, 9999, 1440]
            dur = durations[(idx - 26) % len(durations)]
            input_data = {"title": f"Test Duration {dur}", "category": "work", "effort": "1", "duration": dur, "action": "duration_bounds"}
        elif idx <= 75:
            name = f"Create Intention Title Char Length - Variation {idx-50}"
            desc = "Verify validation for title length (very short, extremely long)."
            steps = "1. Enter varying title lengths\n2. Try to submit"
            expected = "Form behaves stable, titles either truncated, validated, or stored successfully"
            titles = ["A", "B"*100, "C"*300]
            title = titles[(idx - 51) % len(titles)]
            input_data = {"title": title, "category": "learning", "effort": "4", "duration": "60", "action": "title_length"}
        else:
            name = f"Create Intention Emoji Category Title - Variation {idx-75}"
            desc = "Verify creation works with emojis and special character content in fields."
            steps = "1. Enter emoji title\n2. Select category and submit"
            expected = "Intention created and renders emoji symbols correctly in list"
            emoji_titles = ["🏋️‍♂️ Gym Session", "📚 Learn Rust 🚀", "🧘 Meditate 15m", "💻 Code review ⚙️"]
            title = emoji_titles[(idx - 76) % len(emoji_titles)]
            input_data = {"title": title, "category": "personal", "effort": "3", "duration": "15", "action": "emoji_inputs"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # 5. Search & Filters: 50 cases (TC301 to TC350)
    for i in range(301, 351):
        case_id = f"TC{i:03d}"
        module = "Search & Filters"
        
        idx = i - 300
        if idx == 1:
            name = "Filter by Category Health"
            desc = "Verify clicking Category filter lists only matching items."
            steps = "1. Select Category filter: health\n2. Verify all listed intentions belong to 'health'"
            expected = "All displayed intentions list 'health' category label"
            input_data = {"filter_type": "category", "filter_value": "health", "action": "filter_category"}
        elif idx <= 15:
            name = f"Search Matching Query - Variation {idx-1}"
            desc = "Verify search bar filters intentions listing based on matching substring."
            steps = "1. Type matching keyword in search input\n2. Verify listed intentions contain keyword"
            expected = "Intention items matching search string are shown, non-matching are hidden"
            keywords = ["Cardio", "Code", "Rust", "Gym", "Learn", "Meditate"]
            keyword = keywords[(idx - 2) % len(keywords)]
            input_data = {"search_query": keyword, "action": "search_match"}
        elif idx <= 30:
            name = f"Search Non-Matching Query - Variation {idx-15}"
            desc = "Verify search bar shows empty state for non-matching inputs."
            steps = "1. Type random non-matching string in search input\n2. Verify empty list state"
            expected = "No items listed, appropriate empty state display visible"
            input_data = {"search_query": f"random_search_{idx}_xyz", "action": "search_no_match"}
        elif idx <= 40:
            name = f"Filter by Status (Completed/Missed) - Variation {idx-30}"
            desc = "Verify filtering intentions by completed/missed status works."
            steps = "1. Click Completed or Missed filter toggle\n2. Verify list displays correct matching subset"
            status = "completed" if idx % 2 == 0 else "missed"
            expected = f"Only items with status '{status}' are displayed in the list"
            input_data = {"filter_type": "status", "filter_value": status, "action": "filter_status"}
        else:
            name = f"Clear Filters Check - Variation {idx-40}"
            desc = "Verify clicking 'Clear Filters' resets search inputs and filters to show all items."
            steps = "1. Set filters and search query\n2. Click clear button\n3. Verify all intentions list resets"
            expected = "Search query input cleared, category filters reset, all items visible"
            input_data = {"action": "clear_filters"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # 6. Insights & Analytics: 50 cases (TC351 to TC400)
    for i in range(351, 401):
        case_id = f"TC{i:03d}"
        module = "Insights & Analytics"
        
        idx = i - 350
        if idx == 1:
            name = "Verify Completion Rate Stats Card"
            desc = "Verify dashboard / insights stats card shows completion rates correctly."
            steps = "1. Navigate to /insights\n2. Check completion rate widget metrics"
            expected = "Completion rate stats card has visible positive percentage"
            input_data = {"widget_id": "completion_rate", "action": "verify_widget"}
        elif idx <= 15:
            name = f"Verify Chart Rendering - Variation {idx-1}"
            desc = "Verify SVG charts (Pie, Line, Bar, Area) exist and are loaded."
            steps = "1. Navigate to /insights\n2. Wait for charts to load\n3. Assert chart SVG elements are present"
            expected = "Charts loaded successfully in DOM, container displays SVG paths"
            chart_types = ["PieChart", "LineChart", "BarChart", "AreaChart"]
            chart = chart_types[(idx - 2) % len(chart_types)]
            input_data = {"chart_type": chart, "action": "verify_chart"}
        elif idx <= 35:
            name = f"Verify Metric Accuracy - Variation {idx-15}"
            desc = "Verify compliance metrics and time aggregates align with dashboard stats."
            steps = "1. Read summary counts from dashboard\n2. Match calculations with stats labels in Insights"
            expected = "Analytics numbers match dashboard intention state numbers precisely"
            input_data = {"metric": "duration_sum", "action": "verify_metrics"}
        else:
            name = f"Verify Responsive Insights View - Variation {idx-35}"
            desc = "Verify insights layout adjusts cleanly without overlap on mobile sizes."
            steps = "1. Resize browser window to mobile width\n2. Verify grid items are responsive"
            expected = "Grid items stack vertically, charts scale properly to viewport width"
            input_data = {"screen_width": 375, "action": "responsive_check"}

        cases.append((case_id, module, name, desc, steps, expected, json.dumps(input_data)))

    # Write cases to excel sheet
    for row_idx, case in enumerate(cases, start=2):
        for col_idx, val in enumerate(case, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Center ID
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Set column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

    # Save workbook
    wb.save(file_path)
    print(f"Successfully generated {len(cases)} test cases at {file_path}")

if __name__ == "__main__":
    generate_test_cases()
