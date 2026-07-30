import os
import sys
import subprocess
import argparse
from datetime import datetime

# Add local folder to sys path
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))

from config import test_config

def main():
    parser = argparse.ArgumentParser(description="GapLogic Selenium Test Automation Suite Runner")
    parser.add_argument("--browser", default=test_config.BROWSER, choices=["chrome", "firefox", "edge"], help="Browser to run tests on")
    parser.add_argument("--headed", action="store_true", help="Run tests in windowed/headed mode (default is headless)")
    parser.add_argument("--workers", default=test_config.PARALLEL_WORKERS, help="Number of parallel workers (pytest-xdist) e.g. 4 or auto")
    parser.add_argument("-k", help="Filter tests by name expression (passed to pytest -k)")
    parser.add_argument("-m", help="Filter tests by marker expression (passed to pytest -m)")
    args = parser.parse_args()

    suite_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(suite_dir, "data", "test_data.xlsx")

    # 1. Generate test data excel if it doesn't exist
    if not os.path.exists(xlsx_path):
        print("[Runner] Generating test case spreadsheet...")
        from data.generate_test_data import generate_test_cases
        generate_test_cases()
    else:
        print(f"[Runner] Found existing test cases spreadsheet: {xlsx_path}")

    # Determine paths for python and pytest in virtual env
    if sys.platform == "win32":
        pytest_bin = os.path.join(suite_dir, "venv", "Scripts", "pytest.exe")
    else:
        pytest_bin = os.path.join(suite_dir, "venv", "bin", "pytest")

    if not os.path.exists(pytest_bin):
        # Fallback to current environment's pytest if venv is not setup or we are running in general env
        pytest_bin = "pytest"

    # Construct pytest command line arguments
    headless = "false" if args.headed else "true"
    
    cmd = [
        pytest_bin,
        "-v",
        f"--browser-type={args.browser}",
        f"--headless-mode={headless}",
        "--tb=short"
    ]

    # Add parallel execution (pytest-xdist)
    if args.workers and args.workers.lower() != "0":
        cmd.extend(["-n", args.workers])

    # Add filtering if provided
    if args.k:
        cmd.extend(["-k", args.k])
    if args.m:
        cmd.extend(["-m", args.m])

    # Target test file
    cmd.append(os.path.join(suite_dir, "tests", "test_suite.py"))

    print(f"\n[Runner] Running Selenium Suite. Command: {' '.join(cmd)}")
    print(f"[Runner] Execution Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("[Runner] Running 400 data-driven test cases...\n")

    start_time = datetime.now()
    # Run the process
    result = subprocess.run(cmd, cwd=suite_dir)
    end_time = datetime.now()

    duration = (end_time - start_time).total_seconds()
    print(f"\n[Runner] Execution Complete. Duration: {duration:.2f} seconds.")
    print(f"[Runner] Exit Code: {result.returncode}")
    print("[Runner] Compile reporting records...")

    # Let's import conftest helper to merge worker results if needed
    # Note: when session ends, master process compiles results from worker JSONs.
    # We will locate the reports directory and find the latest report
    reports_dir = os.path.join(suite_dir, "reports")
    latest_report = os.path.join(reports_dir, "test_report_latest.xlsx")
    
    if os.path.exists(latest_report):
        print(f"\n========================================================")
        print(f" SUCCESS: Formatted Excel Report Generated Successfully! ")
        print(f" Path: {latest_report}")
        print(f"========================================================\n")
        
        # Display short stats
        try:
            import openpyxl
            wb = openpyxl.load_workbook(latest_report)
            ws = wb["Summary"]
            print("--- EXECUTION REPORT SUMMARY ---")
            for r in range(5, 11):
                label = ws.cell(row=r, column=1).value
                val = ws.cell(row=r, column=2).value
                print(f" {label}: {val}")
            print("--------------------------------")
        except Exception as e:
            print(f"[Runner] Summary could not be printed: {str(e)}")
    else:
        print("[Runner] Warning: Latest Excel report was not found. Check if the test ran and conftest session finish executed.")

    sys.exit(result.returncode)

if __name__ == "__main__":
    main()
