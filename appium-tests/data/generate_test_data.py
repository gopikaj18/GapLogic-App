import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_file = os.path.join(script_dir, "test_data.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ["Test Case ID", "Module/Category", "Test Case Name", "Description", "Steps", "Expected Result", "Input Data"]
    ws.row_dimensions[1].height = 25
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Categories distribution:
    # 1. Onboarding/Splash (60 cases: 1-60)
    # 2. Login/Signup/Auth (100 cases: 61-160)
    # 3. Home/Dashboard/Modeler (100 cases: 161-260)
    # 4. Navigation (60 cases: 261-320)
    # 5. Search & Filters (40 cases: 321-360)
    # 6. Gestures & Device (40 cases: 361-400)

    row_idx = 2
    for i in range(1, 401):
        case_id = f"TC{i:03d}"
        input_data = {}
        
        if i <= 60:
            module = "Splash & Onboarding"
            name = f"Splash Loading State - Var {i}"
            desc = "Verify app splash screen loader, layout rendering, user-agent tag and offline network retry triggers."
            steps = "1. Launch mobile app\n2. Observe Splash screen loader\n3. Wait for layout redirect"
            expected = "Splash screen displays loading spinner, custom UA 'GapLogicAndroid' sent, redirects successfully to auth/dashboard."
            input_data = {"action": "splash_check", "variation": i, "network_connected": True}
            
        elif i <= 160:
            idx = i - 60
            module = "Auth Screen"
            if idx <= 30:
                name = f"Login Validation - Var {idx}"
                desc = "Verify login input behaviors, invalid email check, empty bounds validation."
                steps = "1. Navigate to Login\n2. Enter credentials\n3. Click Submit"
                expected = "Rejects invalid credentials or logs in successfully, user session is initialized."
                input_data = {"action": "login_validation", "variation": idx, "email": f"mobile_user_{idx}@example.com", "password": "password123"}
            elif idx <= 70:
                name = f"Signup Fields - Var {idx-30}"
                desc = "Verify registration field checks, duplicate email checks, and edge case password sizes."
                steps = "1. Navigate to Register\n2. Enter name, email, and password\n3. Click Create Account"
                expected = "User registration completes successfully or appropriate error boundaries are highlighted."
                input_data = {"action": "register_validation", "variation": idx-30, "name": f"Mobile Dev {idx}", "email": f"reg_user_{idx}@example.com", "password": "pass", "confirmPassword": "pass"}
            else:
                name = f"Auth Navigation Redirects - Var {idx-70}"
                desc = "Verify that unauthenticated users are forced back to auth screen when hitting deep links."
                steps = "1. Navigate to protected screen\n2. Inspect redirect layout"
                expected = "User session checked, user automatically routed to Auth Screen."
                input_data = {"action": "auth_redirect", "variation": idx-70, "target_screen": "/modeler"}
                
        elif i <= 260:
            idx = i - 160
            module = "Modeler Screen"
            if idx <= 50:
                name = f"Create Intention Success - Var {idx}"
                desc = "Verify mobile intention creation form works with emoji titles, standard category select, and duration sliders."
                steps = "1. Open Modeler form\n2. Enter title, pick category and duration\n3. Tap Establish Intention"
                expected = "Intention saved successfully, active intention renders correctly on the list panel."
                input_data = {"action": "create_intention", "variation": idx, "title": f"🧘 Meditate {idx}m", "category": "personal", "duration": idx + 5}
            else:
                name = f"Form Field Constraints - Var {idx-50}"
                desc = "Verify that empty titles, negative duration values, or character limits are blocked."
                steps = "1. Open Modeler form\n2. Input out of bound constraints\n3. Tap Establish Intention"
                expected = "Form inputs fail validation, UI display blocks submit action or prompts warning."
                input_data = {"action": "intention_bounds", "variation": idx-50, "title": "", "category": "invalid", "duration": -5}
                
        elif i <= 320:
            idx = i - 260
            module = "Navigation Screen"
            name = f"Tab Layout Switches - Var {idx}"
            desc = "Verify sidebar tabs, active buttons highlights, mobile tab bar responsiveness, and back button routing."
            steps = "1. Tap tab items sequentially\n2. Verify highlighted class list\n3. Tap hardware back button"
            expected = "Screens transition smoothly, selected tab is highlighted, hardware back button rolls back screen history."
            input_data = {"action": "nav_switches", "variation": idx, "target_tab": "Insights"}
            
        elif i <= 360:
            idx = i - 320
            module = "Search & Filter"
            name = f"Dashboard Filter Check - Var {idx}"
            desc = "Verify intention searching, status filtering (Active/Completed/All), and clear filters button."
            steps = "1. Enter search keyword\n2. Select status filter\n3. Click clear filters"
            expected = "Items in the card grid are correctly matched and filtered, list updates cleanly."
            input_data = {"action": "search_filter", "variation": idx, "keyword": "Rust", "status": "completed"}
            
        else:
            idx = i - 360
            module = "Gestures & Device"
            name = f"Gesture Interaction - Var {idx}"
            desc = "Verify horizontal card swipes to complete/delete, vertical list scrolling, and screen rotation layout resizing."
            steps = "1. Swipe intention card horizontally\n2. Scroll down on list\n3. Rotate device to landscape"
            expected = "Intention completed or deleted via swipe, list scrolls smoothly, layouts adapt to screen orientation changes."
            input_data = {"action": "device_gestures", "variation": idx, "gesture": "swipe_right", "rotation": "landscape"}
            
        row_data = [case_id, module, name, desc, steps, expected, json.dumps(input_data)]
        
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 38)
        
    wb.save(xlsx_file)
    print(f"Generated 400 Mobile E2E Test Cases in: {xlsx_file}")

if __name__ == "__main__":
    main()
