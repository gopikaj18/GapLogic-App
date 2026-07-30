import os
import openpyxl
import json
import logging

# Configure logger
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("API_QA")

def load_test_cases_from_excel():
    """
    Parses definitions from data/test_data.xlsx.
    """
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx_path = os.path.join(script_dir, "data", "test_data.xlsx")
    
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Definitions sheet not found at: {xlsx_path}")
        
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Test Cases"]
    
    cases = []
    # Read starting from row 2 (skipping header)
    for row in range(2, ws.max_row + 1):
        case_id = ws.cell(row=row, column=1).value
        if not case_id:
            continue
        cases.append({
            "case_id": case_id,
            "module": ws.cell(row=row, column=2).value,
            "name": ws.cell(row=row, column=3).value,
            "description": ws.cell(row=row, column=4).value,
            "steps": ws.cell(row=row, column=5).value,
            "expected": ws.cell(row=row, column=6).value,
            "input_data": ws.cell(row=row, column=7).value
        })
    return cases
