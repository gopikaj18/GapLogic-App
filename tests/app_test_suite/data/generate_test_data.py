import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_file = os.path.join(script_dir, "test_data.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ["Test Case ID", "Module/Category", "Test Case Name", "Description", "Steps", "Expected Result", "Input Data"]
    ws.row_dimensions[1].height = 25
    
    font_family = "Segoe UI"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_idx = 2
    for i in range(1, 401):
        case_id = f"TC{i:03d}"
        input_data = {}
        
        if i <= 60:
            module = "Login & Session Handling"
            name = f"Auth Validation - Var {i}"
            desc = "Verify backend rejection for invalid logins, empty credentials, and token structures."
            steps = "1. Send POST to /api/auth/login\n2. Provide malformed email or empty payload\n3. Verify HTTP response code is 4xx"
            expected = "Rejects login attempt and returns 4xx Client Error status."
            input_data = {"action": "auth_check", "variation": i, "email": f"invalid_user_{i}@invalid", "password": ""}
            
        elif i <= 140:
            idx = i - 60
            module = "Access Control Correctness"
            name = f"Endpoint Access Check - Var {idx}"
            desc = "Verify protected route redirects or rejects requests missing active local storage session tokens."
            steps = "1. Send GET to protected endpoint\n2. Do not provide Authorization header\n3. Verify redirect to /login or HTTP 401"
            expected = "Access denied. Returns 401 Unauthorized or redirects back to auth screen."
            input_data = {"action": "access_control_check", "variation": idx, "target_endpoint": f"/api/intentions?var={idx}"}
            
        elif i <= 240:
            idx = i - 140
            module = "Input Handling & Validation"
            name = f"Intention Form Validation - Var {idx}"
            desc = "Verify database constraints check, title boundary handling, category selects, and duration sliders."
            steps = "1. Send POST to /api/intentions\n2. Enter boundary values (e.g. empty or extreme integers)\n3. Verify validation error"
            expected = "Input validation blocks transaction. Returns 400 Bad Request."
            input_data = {"action": "input_validation_check", "variation": idx, "title": "", "category": "invalid_cat", "duration": -10}
            
        elif i <= 300:
            idx = i - 240
            module = "API Contract & Headers"
            name = f"HTTP Header Policy - Var {idx}"
            desc = "Verify API response headers (X-Frame-Options, Content-Type, CORS configuration)."
            steps = "1. Send GET to target endpoint\n2. Inspect response headers list\n3. Check headers match security compliance policies"
            expected = "Security headers present, Content-Type matches JSON, CORS matches allowed origins."
            input_data = {"action": "header_check", "variation": idx, "endpoint": "/"}
            
        elif i <= 350:
            idx = i - 300
            module = "Error Containment"
            name = f"Database Error Leak Check - Var {idx}"
            desc = "Verify database validation exceptions do not leak stack traces or directory paths."
            steps = "1. Trigger database validation constraints\n2. Inspect body content\n3. Verify no stack trace is leaked"
            expected = "No database info is leaked. Response body contains simple sanitized error messages."
            input_data = {"action": "error_containment_check", "variation": idx, "payload": "' OR 1=1 --"}
            
        else:
            idx = i - 350
            module = "Throttling & Abuse Prevention"
            name = f"Rate Limit Switch - Var {idx}"
            desc = "Verify rate-limiting checks under repeat concurrent requests."
            steps = "1. Send multiple concurrent requests sequentially\n2. Verify rate-limiting triggers\n3. Check status code returns 429 Too Many Requests"
            expected = "API throttling triggers successfully. Client gets 429 status code."
            input_data = {"action": "throttling_check", "variation": idx, "burst_count": idx + 5}
            
        row_data = [case_id, module, name, desc, steps, expected, json.dumps(input_data)]
        
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 38)
        
    wb.save(xlsx_file)
    print(f"Generated 400 API Robustness QA cases in: {xlsx_file}")

if __name__ == "__main__":
    main()
