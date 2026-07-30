import os
import random
import string
import openpyxl
import logging
from datetime import datetime
from config.test_config import SCREENSHOT_DIR

def setup_logger(name="selenium_tests"):
    logger = logging.getLogger(name)
    if not logger.handlers:
        logger.setLevel(logging.INFO)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        ch.setFormatter(formatter)
        logger.addHandler(ch)
    return logger

logger = setup_logger()

def load_test_cases_from_excel():
    data_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(data_dir, "data", "test_data.xlsx")
    
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Test data file not found at {file_path}. Please run generate_test_data.py first.")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    test_cases = []
    # Skip header
    for row in range(2, ws.max_row + 1):
        case_id = ws.cell(row=row, column=1).value
        module = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value
        desc = ws.cell(row=row, column=4).value
        steps = ws.cell(row=row, column=5).value
        expected = ws.cell(row=row, column=6).value
        input_data = ws.cell(row=row, column=7).value

        if case_id:
            test_cases.append({
                "case_id": str(case_id),
                "module": str(module),
                "name": str(name),
                "description": str(desc),
                "steps": str(steps),
                "expected": str(expected),
                "input_data": str(input_data)
            })

    return test_cases

def generate_random_email():
    chars = string.ascii_lowercase + string.digits
    rand_str = ''.join(random.choices(chars, k=8))
    return f"test_{rand_str}@example.com"

def generate_random_name():
    first_names = ["Alex", "Jordan", "Taylor", "Morgan", "Sam", "Chris"]
    last_names = ["Smith", "Jones", "Miller", "Davis", "Garcia", "Rodriguez"]
    return f"{random.choice(first_names)} {random.choice(last_names)}"

def capture_failure_screenshot(driver, case_id):
    if not os.path.exists(SCREENSHOT_DIR):
        os.makedirs(SCREENSHOT_DIR)
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{case_id}_{timestamp}.png"
    filepath = os.path.join(SCREENSHOT_DIR, filename)
    
    try:
        driver.save_screenshot(filepath)
        logger.info(f"Screenshot saved for {case_id} at {filepath}")
        return filepath
    except Exception as e:
        logger.error(f"Failed to capture screenshot for {case_id}: {str(e)}")
        return ""
